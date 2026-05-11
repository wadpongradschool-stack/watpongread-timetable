"""
ระบบจัดตารางเรียนตารางสอนอัตโนมัติ
โรงเรียนวัดโป่งแรด — v3.0
"""

import streamlit as st
import pandas as pd
import random
from copy import deepcopy
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="ตารางสอน | วัดโป่งแรด",
    page_icon="🏫",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sarabun:wght@300;400;600;700;800&display=swap');
* { font-family: 'Sarabun', sans-serif !important; }
.main-header {
    background: linear-gradient(135deg,#1a472a 0%,#2d6a4f 55%,#52b788 100%);
    padding:1.5rem 2rem; border-radius:14px; margin-bottom:1.4rem;
    box-shadow:0 6px 24px rgba(26,71,42,.28);
}
.main-header h1{color:#fff;font-size:1.75rem;font-weight:800;margin:0;}
.main-header p {color:#b7e4c7;font-size:.9rem;margin:.2rem 0 0;}
.stat-card{
    background:#fff;border-radius:12px;padding:.9rem 1.3rem;
    border-left:5px solid #40916c;
    box-shadow:0 2px 10px rgba(0,0,0,.07);margin-bottom:.7rem;
}
.stat-number{font-size:1.9rem;font-weight:700;color:#1a472a;line-height:1;}
.stat-label {font-size:.82rem;color:#6c757d;margin-top:.18rem;}
.stButton>button{
    background:linear-gradient(135deg,#1a472a,#40916c);
    color:#fff;border:none;border-radius:9px;
    padding:.5rem 1.3rem;font-weight:700;font-size:.97rem;
    box-shadow:0 3px 10px rgba(26,71,42,.22);transition:all .18s;
}
.stButton>button:hover{transform:translateY(-2px);box-shadow:0 5px 16px rgba(26,71,42,.35);}
.stTabs [data-baseweb="tab-list"]{
    gap:4px;background:#f0f4f1;border-radius:10px;padding:4px;
}
.stTabs [data-baseweb="tab"]{border-radius:8px;font-weight:600;}
.stTabs [aria-selected="true"]{background:#1a472a !important;color:#fff !important;}
</style>
""", unsafe_allow_html=True)

DAYS = ["จันทร์","อังคาร","พุธ","พฤหัสบดี","ศุกร์"]
PERIOD_TIMES = {
    1:"08:30 – 09:30",
    2:"09:30 – 10:30",
    3:"10:30 – 11:30",
    4:"12:30 – 13:30",
    5:"13:30 – 14:30",
    6:"14:30 – 15:30",
}
PERIODS = list(PERIOD_TIMES.keys())
MORNING = [1, 2, 3]
LUNCH_DISPLAY = "11:30 – 12:30  🍱  พักกลางวัน"
CAT_CORE="วิชาหลัก"; CAT_ELECTIVE="วิชาเสริม"; CAT_ACTIVITY="กิจกรรม"
CAT_OPTIONS=[CAT_CORE,CAT_ELECTIVE,CAT_ACTIVITY]
ROOM_OPTIONS=["ห้องเรียน","ห้องคอมพิวเตอร์","สนามกีฬา","ห้องดนตรี","ห้องศิลปะ","อาคารอเนกประสงค์"]

DEFAULT_TEACHERS=[
    {"id":1,"name":"น.ส.นภาวรรณ สันดิษฐ์",    "homeroom":"อนุบาล 2-3","specialty":"ครูประจำชั้น อนุบาล 2-3"},
    {"id":2,"name":"น.ส.ธารารัตน์ สามิภักดิ์", "homeroom":"อนุบาล 1",  "specialty":"ครูประจำชั้น อนุบาล 1"},
    {"id":3,"name":"นายณัฐวุฒิ คำจันทร์",       "homeroom":"ป.5",       "specialty":"สุขศึกษาและพลศึกษา"},
    {"id":4,"name":"น.ส.ชมพูนุท บุญอากาศ",      "homeroom":"ป.6",       "specialty":"ภาษาอังกฤษ"},
    {"id":5,"name":"น.ส.มณีมัญช์ ศาสตร์ทรัพย์", "homeroom":"ป.4",       "specialty":"วิทยาศาสตร์"},
    {"id":6,"name":"น.ส.กชกร อรัญวัชน์",        "homeroom":"ป.2",       "specialty":"ภาษาไทย"},
    {"id":7,"name":"น.ส.รักชนก โสภักต์",        "homeroom":"ป.3",       "specialty":"วิทยาการคำนวณ"},
    {"id":8,"name":"น.ส.ณิชนันทน์ การบุญ",      "homeroom":"ป.1",       "specialty":"ทุกวิชา (ป.1)"},
]
DEFAULT_CLASSES=["อนุบาล 1","อนุบาล 2","อนุบาล 3","ป.1","ป.2","ป.3","ป.4","ป.5","ป.6"]
DEFAULT_SUBJECTS=[
    {"id":1, "name":"ภาษาไทย",            "category":CAT_CORE,    "periods":5,"needs_focus":True, "room":"ห้องเรียน",        "color":"#d8f3dc"},
    {"id":2, "name":"คณิตศาสตร์",         "category":CAT_CORE,    "periods":5,"needs_focus":True, "room":"ห้องเรียน",        "color":"#caf0f8"},
    {"id":3, "name":"วิทยาศาสตร์",        "category":CAT_CORE,    "periods":3,"needs_focus":True, "room":"ห้องเรียน",        "color":"#e2d9f3"},
    {"id":4, "name":"ภาษาอังกฤษ",         "category":CAT_CORE,    "periods":4,"needs_focus":True, "room":"ห้องเรียน",        "color":"#ffd6a5"},
    {"id":5, "name":"สังคมศึกษา",         "category":CAT_CORE,    "periods":3,"needs_focus":False,"room":"ห้องเรียน",        "color":"#ffccd5"},
    {"id":6, "name":"ภาษาอังกฤษเพิ่มเติม","category":CAT_ELECTIVE,"periods":2,"needs_focus":True, "room":"ห้องเรียน",        "color":"#ffe5b4"},
    {"id":7, "name":"วิทยาการคำนวณ",      "category":CAT_ELECTIVE,"periods":2,"needs_focus":True, "room":"ห้องคอมพิวเตอร์", "color":"#c8e6c9"},
    {"id":8, "name":"สุขศึกษา",           "category":CAT_ELECTIVE,"periods":1,"needs_focus":False,"room":"ห้องเรียน",        "color":"#b2dfdb"},
    {"id":9, "name":"พลศึกษา",            "category":CAT_ELECTIVE,"periods":2,"needs_focus":False,"room":"สนามกีฬา",         "color":"#dcedc8"},
    {"id":10,"name":"ศิลปะ",              "category":CAT_ELECTIVE,"periods":2,"needs_focus":False,"room":"ห้องเรียน",        "color":"#f8bbd0"},
    {"id":11,"name":"ดนตรี",              "category":CAT_ELECTIVE,"periods":1,"needs_focus":False,"room":"ห้องดนตรี",        "color":"#e1bee7"},
    {"id":12,"name":"กิจกรรมชุมนุม",      "category":CAT_ACTIVITY,"periods":1,"needs_focus":False,"room":"ห้องเรียน",        "color":"#fff9c4"},
    {"id":13,"name":"แนะแนว",             "category":CAT_ACTIVITY,"periods":1,"needs_focus":False,"room":"ห้องเรียน",        "color":"#e2e3e5"},
    {"id":14,"name":"ลูกเสือ/เนตรนารี",   "category":CAT_ACTIVITY,"periods":1,"needs_focus":False,"room":"สนามกีฬา",         "color":"#ffe0b2"},
    {"id":15,"name":"ประชุมนักเรียน",      "category":CAT_ACTIVITY,"periods":1,"needs_focus":False,"room":"อาคารอเนกประสงค์","color":"#fce4ec"},
]
DEFAULT_TEACHER_SUBJECTS={
    1:["ภาษาไทย","คณิตศาสตร์","สังคมศึกษา","ศิลปะ","แนะแนว","ประชุมนักเรียน"],
    2:["ภาษาไทย","คณิตศาสตร์","สังคมศึกษา","ศิลปะ","แนะแนว","ประชุมนักเรียน"],
    3:["สุขศึกษา","พลศึกษา","ลูกเสือ/เนตรนารี","กิจกรรมชุมนุม"],
    4:["ภาษาอังกฤษ","ภาษาอังกฤษเพิ่มเติม","แนะแนว"],
    5:["วิทยาศาสตร์","กิจกรรมชุมนุม","แนะแนว"],
    6:["ภาษาไทย","สังคมศึกษา","แนะแนว"],
    7:["วิทยาการคำนวณ","คณิตศาสตร์","กิจกรรมชุมนุม"],
    8:["ภาษาไทย","คณิตศาสตร์","วิทยาศาสตร์","ภาษาอังกฤษ",
       "สังคมศึกษา","ศิลปะ","พลศึกษา","สุขศึกษา","ดนตรี","แนะแนว"],
}
DEFAULT_CLASS_HOMEROOM={
    "อนุบาล 1":2,"อนุบาล 2":1,"อนุบาล 3":1,
    "ป.1":8,"ป.2":6,"ป.3":7,"ป.4":5,"ป.5":3,"ป.6":4,
}

def _init():
    defaults={
        "teachers":deepcopy(DEFAULT_TEACHERS),
        "subjects":deepcopy(DEFAULT_SUBJECTS),
        "classes":list(DEFAULT_CLASSES),
        "teacher_subjects":deepcopy(DEFAULT_TEACHER_SUBJECTS),
        "class_homeroom":deepcopy(DEFAULT_CLASS_HOMEROOM),
        "schedule":None,"teacher_schedules":{},"conflicts":[],"t_by_id":{},
    }
    for k,v in defaults.items():
        if k not in st.session_state: st.session_state[k]=v
_init()

def build_color_map():
    m={s["name"]:s["color"] for s in st.session_state.subjects}
    m["—"]="#f8f9fa"; m["พักกลางวัน"]="#fff3cd"
    return m

def generate_timetable(soft_focus,soft_gap):
    random.seed(42)
    classes=st.session_state.classes; subjects=st.session_state.subjects
    teachers=st.session_state.teachers; ts_map=st.session_state.teacher_subjects
    homeroom=st.session_state.class_homeroom
    t_by_id={t["id"]:t for t in teachers}
    schedule={cls:{day:{p:None for p in PERIODS} for day in DAYS} for cls in classes}
    teacher_busy={t["id"]:{d:{p:False for p in PERIODS} for d in DAYS} for t in teachers}
    tasks=[]
    for cls in classes:
        hid=homeroom.get(cls)
        for subj in subjects:
            sname=subj["name"]
            eligible=[t["id"] for t in teachers if sname in ts_map.get(t["id"],[])]
            if not eligible: eligible=[hid] if hid else [teachers[0]["id"]]
            tid=eligible[0]
            for _ in range(subj["periods"]):
                tasks.append({"class":cls,"subject":subj,"teacher_id":tid})
    tasks.sort(key=lambda x:(0 if x["subject"]["needs_focus"] else 1,-x["subject"]["periods"]))
    conflicts=[]
    for task in tasks:
        cls,subj,tid=task["class"],task["subject"],task["teacher_id"]
        candidates=[]
        for day in DAYS:
            for p in PERIODS:
                if schedule[cls][day][p]: continue
                if teacher_busy[tid][day][p]: continue
                score=0
                if soft_focus and subj["needs_focus"] and p in MORNING: score+=10
                if soft_gap:
                    used=sum(1 for pp in PERIODS if teacher_busy[tid][day][pp])
                    if used<len(PERIODS)-1: score+=3
                candidates.append((score,random.random(),day,p))
        if candidates:
            candidates.sort(key=lambda x:(-x[0],x[1]))
            _,_,bd,bp=candidates[0]
            schedule[cls][bd][bp]={
                "subject":subj["name"],"teacher_id":tid,
                "room":subj["room"],"category":subj["category"],"color":subj["color"],
            }
            teacher_busy[tid][bd][bp]=True
        else:
            conflicts.append(f"{cls} — {subj['name']}")
    return schedule,conflicts,t_by_id

def build_teacher_sched(tid,schedule):
    result={d:{p:None for p in PERIODS} for d in DAYS}
    for cls in st.session_state.classes:
        for day in DAYS:
            for p in PERIODS:
                slot=schedule[cls][day][p]
                if slot and slot["teacher_id"]==tid:
                    result[day][p]={"class":cls,"subject":slot["subject"],"color":slot.get("color","#fff")}
    return result

def export_excel():
    schedule=st.session_state.schedule
    teacher_schedules=st.session_state.teacher_schedules
    teachers=st.session_state.teachers; classes=st.session_state.classes
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    H_FILL=PatternFill("solid",fgColor="1A472A"); S_FILL=PatternFill("solid",fgColor="40916C")
    L_FILL=PatternFill("solid",fgColor="FFF3CD"); F_FILL=PatternFill("solid",fgColor="F8F9FA")
    G_FILL=PatternFill("solid",fgColor="E9ECEF"); A_FILL=PatternFill("solid",fgColor="F0FAF4")
    HFONT=Font(name="TH Sarabun New",bold=True,color="FFFFFF",size=13)
    BFONT=Font(name="TH Sarabun New",size=11)
    GFONT=Font(name="TH Sarabun New",bold=True,size=11)
    TFONT=Font(name="TH Sarabun New",bold=True,color="FFFFFF",size=15)
    LFONT=Font(name="TH Sarabun New",bold=True,color="856404",size=12)
    CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
    BRD=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    def sc(cell,fill=F_FILL,font=BFONT): cell.fill=fill;cell.font=font;cell.alignment=CTR;cell.border=BRD
    def hx(c): return c.lstrip("#") if c and c.startswith("#") else "FFFFFF"
    def write_sheet(ws,title_text,row_fn):
        ws.sheet_view.showGridLines=False
        n_cols=len(DAYS)+1; last_col=get_column_letter(n_cols)
        ws.merge_cells(f"A1:{last_col}1"); ws["A1"]=title_text
        ws["A1"].fill=H_FILL;ws["A1"].font=TFONT;ws["A1"].alignment=CTR;ws.row_dimensions[1].height=36
        ws.column_dimensions["A"].width=22
        sc(ws.cell(row=2,column=1,value="เวลา / คาบ"),fill=S_FILL,font=HFONT)
        for ci,day in enumerate(DAYS,2):
            sc(ws.cell(row=2,column=ci,value=day),fill=S_FILL,font=HFONT)
            ws.column_dimensions[get_column_letter(ci)].width=22
        ws.row_dimensions[2].height=28
        row=3
        for p in PERIODS:
            if p==4:
                ws.merge_cells(f"A{row}:{last_col}{row}")
                ws[f"A{row}"]="11:30 – 12:30   🍱  พักกลางวัน"
                ws[f"A{row}"].fill=L_FILL;ws[f"A{row}"].font=LFONT;ws[f"A{row}"].alignment=CTR
                ws.row_dimensions[row].height=26;row+=1
            label=f"คาบ {p}\n{PERIOD_TIMES[p]}"
            sc(ws.cell(row=row,column=1,value=label),fill=G_FILL,font=GFONT)
            ws.row_dimensions[row].height=42;row_fn(ws,row,p);row+=1
    for cls in classes:
        ws=wb.create_sheet(title=cls[:28])
        def rfn(ws,row,p,_cls=cls):
            for ci,day in enumerate(DAYS,2):
                slot=schedule[_cls][day][p]
                if slot:
                    t_obj=next((t for t in teachers if t["id"]==slot["teacher_id"]),None)
                    tn=t_obj["name"].split()[-1] if t_obj else ""
                    text=f"{slot['subject']}\n({tn})"
                    fill=PatternFill("solid",fgColor=hx(slot.get("color","#ffffff")))
                else: text="—";fill=F_FILL
                sc(ws.cell(row=row,column=ci,value=text),fill=fill)
        write_sheet(ws,f"ตารางเรียน ชั้น {cls}  —  โรงเรียนวัดโป่งแรด",rfn)
    t_by_id={t["id"]:t for t in teachers}
    for tid,ts in teacher_schedules.items():
        tname=t_by_id[tid]["name"]
        ws=wb.create_sheet(title=tname.split()[-1][:28])
        def rfn(ws,row,p,_ts=ts):
            for ci,day in enumerate(DAYS,2):
                slot=_ts.get(day,{}).get(p)
                if slot:
                    text=f"{slot['subject']}\n({slot['class']})"
                    fill=PatternFill("solid",fgColor=hx(slot.get("color","#ffffff")))
                else: text="— ว่าง —";fill=F_FILL
                sc(ws.cell(row=row,column=ci,value=text),fill=fill)
        write_sheet(ws,f"ตารางสอน — {tname}  |  โรงเรียนวัดโป่งแรด",rfn)
    ws=wb.create_sheet(title="สรุปภาระงาน",index=0)
    ws.sheet_view.showGridLines=False
    hdrs=["ชื่อครู","ห้องประจำ","วิชาหลัก/ความชำนาญ","คาบสอน/สัปดาห์","คาบว่าง","สถานะ"]
    wths=[30,16,26,20,12,14]
    ws.merge_cells(f"A1:{get_column_letter(len(hdrs))}1")
    ws["A1"]="สรุปภาระงานสอน — โรงเรียนวัดโป่งแรด"
    ws["A1"].fill=H_FILL;ws["A1"].font=TFONT;ws["A1"].alignment=CTR;ws.row_dimensions[1].height=40
    for ci,(h,w) in enumerate(zip(hdrs,wths),1):
        sc(ws.cell(row=2,column=ci,value=h),fill=S_FILL,font=HFONT)
        ws.column_dimensions[get_column_letter(ci)].width=w
    for ri,t in enumerate(teachers,3):
        ts=teacher_schedules.get(t["id"],{})
        used=sum(1 for d in DAYS for p in PERIODS if ts.get(d,{}).get(p))
        free=len(DAYS)*len(PERIODS)-used
        status="✅ ปกติ" if free>=5 else "⚠️ งานหนัก"
        fill=A_FILL if ri%2==0 else F_FILL
        for ci,val in enumerate([t["name"],t["homeroom"],t["specialty"],used,free,status],1):
            sc(ws.cell(row=ri,column=ci,value=val),fill=fill)
        ws.row_dimensions[ri].height=26
    buf=BytesIO();wb.save(buf);buf.seek(0)
    return buf

# SIDEBAR
with st.sidebar:
    st.markdown("## ⚙️ ตั้งค่าระบบ")
    st.markdown("---")
    st.markdown("### 🔧 Soft Constraints")
    soft_focus=st.toggle("วิชา Focus ช่วงเช้า (คาบ 1-3)",value=True)
    soft_gap=st.toggle("ครูมีคาบว่างอย่างน้อย 1 คาบ/วัน",value=True)
    st.markdown("---")
    n_t=len(st.session_state.teachers); n_s=len(st.session_state.subjects)
    n_c=len(st.session_state.classes);  n_cf=len(st.session_state.conflicts)
    st.info(f"👩‍🏫 ครู **{n_t}** คน  |  📚 วิชา **{n_s}**  |  🏠 **{n_c}** ชั้น")
    if st.session_state.schedule:
        if n_cf: st.warning(f"⚠️ Conflict: {n_cf} รายการ")
        else:    st.success("✅ ไม่มี Conflict")
    st.markdown("---")
    if st.button("🚀 สร้างตารางอัตโนมัติ",use_container_width=True):
        with st.spinner("กำลังประมวลผล..."):
            s,cf,tbi=generate_timetable(soft_focus,soft_gap)
            st.session_state.schedule=s; st.session_state.conflicts=cf; st.session_state.t_by_id=tbi
            st.session_state.teacher_schedules={
                t["id"]:build_teacher_sched(t["id"],s) for t in st.session_state.teachers
            }
        st.success("✅ สร้างตารางเสร็จแล้ว!"); st.rerun()
    if st.session_state.schedule:
        buf=export_excel()
        st.download_button("📥 Export Excel",data=buf,
            file_name="ตารางสอน_วัดโป่งแรด.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

# HEADER
st.markdown("""
<div class="main-header">
  <h1>🏫 ระบบจัดตารางเรียนตารางสอนอัตโนมัติ</h1>
  <p>โรงเรียนวัดโป่งแรด · ปีการศึกษา 2568</p>
</div>
""",unsafe_allow_html=True)

c1,c2,c3,c4=st.columns(4)
for col,num,label,accent in [
    (c1,n_t,"👩‍🏫 ครูทั้งหมด","#1a472a"),(c2,n_s,"📚 วิชาทั้งหมด","#1a472a"),
    (c3,n_c,"🏠 ชั้นเรียน","#1a472a"),(c4,n_cf,"⚠️ Conflicts","#dc3545" if n_cf else "#1a472a"),
]:
    col.markdown(f'<div class="stat-card" style="border-left-color:{accent}"><div class="stat-number" style="color:{accent}">{num}</div><div class="stat-label">{label}</div></div>',unsafe_allow_html=True)

tab1,tab2,tab3,tab4,tab5,tab6=st.tabs([
    "📊 ตารางเรียนรายชั้น","👩‍🏫 ตารางสอนรายครู",
    "✏️ จัดการครู","📚 จัดการวิชา","🔗 มอบหมายวิชา→ครู","🔍 ตรวจสอบผล",
])

cmap=build_color_map()

def make_rows(period_fn):
    rows=[]
    for p in PERIODS:
        if p==4:
            row_lunch={"เวลา / คาบ":f"🍱  {LUNCH_DISPLAY}"}
            for d in DAYS: row_lunch[d]="พักกลางวัน"
            rows.append(row_lunch)
        row={"เวลา / คาบ":f"คาบ {p}  ·  {PERIOD_TIMES[p]}"}
        for d in DAYS: row[d]=period_fn(d,p)
        rows.append(row)
    return rows

def render_timetable(rows,color_fn):
    df=pd.DataFrame(rows).set_index("เวลา / คาบ")
    st.dataframe(df.style.map(color_fn),use_container_width=True,height=440)

# TAB 1
with tab1:
    if not st.session_state.schedule:
        st.info("👈 กด **🚀 สร้างตารางอัตโนมัติ** ที่แถบซ้ายเพื่อเริ่มต้น")
    else:
        selected_cls=st.selectbox("เลือกชั้นเรียน",st.session_state.classes,key="cls_sel")
        sched=st.session_state.schedule[selected_cls]
        def pf1(d,p):
            slot=sched[d][p]
            if not slot: return "—"
            t=st.session_state.t_by_id.get(slot["teacher_id"])
            tn=t["name"].split()[-1] if t else ""
            return f"{slot['subject']} ({tn})"
        def cf1(val):
            if "พักกลางวัน" in str(val): return "background-color:#fff3cd;color:#856404;font-weight:700"
            subj=val.split(" (")[0] if "(" in val else val
            return f"background-color:{cmap.get(subj,'#f8f9fa')};font-weight:500"
        st.markdown(f"#### ตารางเรียน ชั้น {selected_cls}")
        render_timetable(make_rows(pf1),cf1)
        st.markdown("**สีวิชา:**")
        cats={}
        for s in st.session_state.subjects: cats.setdefault(s["category"],[]).append(s)
        for cat,subjs in cats.items():
            badges="  ".join(f'<span style="background:{s["color"]};padding:2px 8px;border-radius:4px;font-size:.8rem">{s["name"]}</span>' for s in subjs)
            st.markdown(f"*{cat}* — {badges}",unsafe_allow_html=True)

# TAB 2
with tab2:
    if not st.session_state.schedule:
        st.info("👈 กด **🚀 สร้างตารางอัตโนมัติ** ก่อนครับ")
    else:
        sel_t=st.selectbox("เลือกครู",[t["name"] for t in st.session_state.teachers],key="t_sel")
        t_obj=next(t for t in st.session_state.teachers if t["name"]==sel_t)
        ts=st.session_state.teacher_schedules.get(t_obj["id"],{})
        used=sum(1 for d in DAYS for p in PERIODS if ts.get(d,{}).get(p))
        free=len(DAYS)*len(PERIODS)-used
        ca,cb,cc=st.columns([3,1,1])
        ca.markdown(f"#### ตารางสอน — {sel_t}")
        ca.caption(f"วิชาหลัก: {t_obj['specialty']}  ·  ประจำชั้น: {t_obj['homeroom']}")
        cb.metric("คาบสอน/สัปดาห์",used); cc.metric("คาบว่าง/สัปดาห์",free)
        def pf2(d,p):
            slot=ts.get(d,{}).get(p)
            return f"{slot['subject']} ({slot['class']})" if slot else "— ว่าง —"
        def cf2(val):
            if "พักกลางวัน" in str(val): return "background-color:#fff3cd;color:#856404;font-weight:700"
            if "ว่าง" in str(val): return "background-color:#f8f9fa;color:#adb5bd"
            subj=val.split(" (")[0]
            return f"background-color:{cmap.get(subj,'#fff')};font-weight:500"
        render_timetable(make_rows(pf2),cf2)

# TAB 3 จัดการครู
with tab3:
    st.markdown("#### 👩‍🏫 รายชื่อครู — เพิ่ม / แก้ไข / ลบ")
    st.caption("กด **+ Add row** เพื่อเพิ่มครู  ·  เลือกแถวแล้วกด 🗑 เพื่อลบ")
    edited_t=st.data_editor(
        pd.DataFrame(st.session_state.teachers),num_rows="dynamic",
        use_container_width=True,hide_index=True,
        column_config={
            "id":st.column_config.NumberColumn("ID",width=55),
            "name":st.column_config.TextColumn("ชื่อ-นามสกุล",width=220),
            "homeroom":st.column_config.TextColumn("ห้องประจำ",width=130),
            "specialty":st.column_config.TextColumn("วิชาหลัก / ความชำนาญ",width=240),
        },
    )
    c3a,c3b=st.columns(2)
    with c3a:
        if st.button("💾 บันทึกข้อมูลครู",use_container_width=True):
            recs=edited_t.to_dict("records")
            for i,r in enumerate(recs):
                if not r.get("id"): r["id"]=i+1
            st.session_state.teachers=recs; st.session_state.schedule=None
            st.success(f"✅ บันทึก {len(recs)} คน — กด สร้างตารางอีกครั้ง")
    with c3b:
        if st.button("🔄 รีเซ็ตข้อมูลครูตั้งต้น",use_container_width=True):
            st.session_state.teachers=deepcopy(DEFAULT_TEACHERS); st.session_state.schedule=None
            st.success("รีเซ็ตเรียบร้อย"); st.rerun()
    st.markdown("---")
    st.markdown("##### ➕ เพิ่มครูด่วน")
    qa1,qa2,qa3=st.columns([2,1,2])
    new_tname=qa1.text_input("ชื่อ-นามสกุล",placeholder="เช่น น.ส.สมหญิง ใจดี",key="nt_name")
    new_homeroom=qa2.text_input("ห้องประจำ",placeholder="เช่น ป.4",key="nt_room")
    new_spec=qa3.text_input("วิชาหลัก/ความชำนาญ",placeholder="เช่น คณิตศาสตร์",key="nt_spec")
    if st.button("➕ เพิ่มครูเลย",key="add_teacher_btn"):
        if new_tname.strip():
            new_id=max((t["id"] for t in st.session_state.teachers),default=0)+1
            st.session_state.teachers.append({"id":new_id,"name":new_tname.strip(),"homeroom":new_homeroom.strip(),"specialty":new_spec.strip()})
            st.session_state.schedule=None; st.success(f"เพิ่มครู '{new_tname}' แล้ว!"); st.rerun()
        else: st.warning("กรุณาใส่ชื่อครู")

# TAB 4 จัดการวิชา
with tab4:
    st.markdown("#### 📚 รายวิชาทั้งหมด — เพิ่ม / แก้ไข / ลบ")
    st.caption("เพิ่มวิชาเสริม กิจกรรม หรือลบวิชาที่ไม่ต้องการ · สีกำหนดได้ (รูปแบบ hex เช่น #d8f3dc)")
    edited_s=st.data_editor(
        pd.DataFrame(st.session_state.subjects),num_rows="dynamic",
        use_container_width=True,hide_index=True,
        column_config={
            "id":st.column_config.NumberColumn("ID",width=52),
            "name":st.column_config.TextColumn("ชื่อวิชา / กิจกรรม",width=190),
            "category":st.column_config.SelectboxColumn("หมวดหมู่",width=120,options=CAT_OPTIONS),
            "periods":st.column_config.NumberColumn("คาบ/สัปดาห์",width=110,min_value=1,max_value=30),
            "needs_focus":st.column_config.CheckboxColumn("ต้องการสมาธิ",width=110),
            "room":st.column_config.SelectboxColumn("ห้อง/สถานที่",width=170,options=ROOM_OPTIONS),
            "color":st.column_config.TextColumn("สี (hex)",width=100),
        },
    )
    c4a,c4b=st.columns(2)
    with c4a:
        if st.button("💾 บันทึกข้อมูลวิชา",use_container_width=True):
            recs=edited_s.to_dict("records")
            for i,r in enumerate(recs):
                if not r.get("id"): r["id"]=i+1
                if not r.get("color"): r["color"]="#e2e3e5"
                if r.get("periods") is None: r["periods"]=1
            st.session_state.subjects=recs; st.session_state.schedule=None
            st.success(f"✅ บันทึก {len(recs)} วิชา — กด สร้างตารางอีกครั้ง")
    with c4b:
        if st.button("🔄 รีเซ็ตวิชาตั้งต้น",use_container_width=True):
            st.session_state.subjects=deepcopy(DEFAULT_SUBJECTS); st.session_state.schedule=None
            st.success("รีเซ็ตเรียบร้อย"); st.rerun()
    st.markdown("---")
    st.markdown("##### ➕ เพิ่มวิชา / กิจกรรมด่วน")
    qa1,qa2,qa3,qa4,qa5=st.columns([2,1,1,1,1])
    new_sname=qa1.text_input("ชื่อวิชา/กิจกรรม",placeholder="เช่น ชมรมหุ่นยนต์",key="ns_name")
    new_cat=qa2.selectbox("หมวด",CAT_OPTIONS,index=2,key="ns_cat")
    new_per=qa3.number_input("คาบ/สัปดาห์",1,30,1,key="ns_per")
    new_room=qa4.selectbox("ห้อง",ROOM_OPTIONS,key="ns_room")
    new_color=qa5.color_picker("สี","#e2e3e5",key="ns_color")
    if st.button("➕ เพิ่มวิชาเลย",key="add_subj_btn"):
        if new_sname.strip():
            new_id=max((s["id"] for s in st.session_state.subjects),default=0)+1
            st.session_state.subjects.append({"id":new_id,"name":new_sname.strip(),"category":new_cat,"periods":int(new_per),"needs_focus":False,"room":new_room,"color":new_color})
            st.session_state.schedule=None; st.success(f"เพิ่ม '{new_sname}' แล้ว!"); st.rerun()
        else: st.warning("กรุณาใส่ชื่อวิชา")

# TAB 5 มอบหมายวิชา→ครู
with tab5:
    st.markdown("#### 🔗 กำหนดว่าครูแต่ละคนสอนวิชาอะไรได้บ้าง")
    st.caption("ติ๊กถูกในช่องที่ต้องการ แล้วกด บันทึก")
    all_subjects=[s["name"] for s in st.session_state.subjects]
    ts_map=st.session_state.teacher_subjects
    rows_ts=[]
    for t in st.session_state.teachers:
        row={"ชื่อครู":t["name"],"ห้องประจำ":t["homeroom"]}
        for sname in all_subjects: row[sname]=sname in ts_map.get(t["id"],[])
        rows_ts.append(row)
    col_cfg={"ชื่อครู":st.column_config.TextColumn(width=200,disabled=True),"ห้องประจำ":st.column_config.TextColumn(width=110,disabled=True)}
    for sname in all_subjects: col_cfg[sname]=st.column_config.CheckboxColumn(sname,width=130)
    edited_ts=st.data_editor(pd.DataFrame(rows_ts),use_container_width=True,hide_index=True,column_config=col_cfg)
    if st.button("💾 บันทึกการมอบหมายวิชา",use_container_width=True):
        new_map={}
        for i,t in enumerate(st.session_state.teachers):
            row=edited_ts.iloc[i]
            new_map[t["id"]]=[s for s in all_subjects if row.get(s,False)]
        st.session_state.teacher_subjects=new_map; st.session_state.schedule=None
        st.success("✅ บันทึกแล้ว — กด สร้างตารางอีกครั้ง")

# TAB 6 ตรวจสอบผล
with tab6:
    st.markdown("#### 🔍 ผลการตรวจสอบ Constraints")
    if not st.session_state.schedule:
        st.info("ยังไม่ได้สร้างตาราง")
    else:
        cfs=st.session_state.conflicts
        if not cfs: st.success("✅ ไม่พบ Conflict! ตารางสอนสมบูรณ์")
        else:
            st.error(f"❌ พบ {len(cfs)} Conflict")
            for c in cfs: st.markdown(f"- ⚠️ {c}")
        st.markdown("---")
        st.markdown("#### 📊 สรุปภาระงานสอนรายครู")
        summary=[]
        for t in st.session_state.teachers:
            ts=st.session_state.teacher_schedules.get(t["id"],{})
            used=sum(1 for d in DAYS for p in PERIODS if ts.get(d,{}).get(p))
            free=len(DAYS)*len(PERIODS)-used
            summary.append({"ชื่อครู":t["name"],"ห้องประจำ":t["homeroom"],"วิชาหลัก":t["specialty"],"คาบสอน/สัปดาห์":used,"คาบว่าง/สัปดาห์":free,"สถานะ":"✅ ปกติ" if free>=5 else "⚠️ งานหนัก"})
        def hl(val):
            if "⚠️" in str(val): return "background-color:#fff3cd;color:#856404;font-weight:600"
            if "✅" in str(val):  return "background-color:#d1e7dd;color:#0f5132;font-weight:600"
            return ""
        st.dataframe(pd.DataFrame(summary).style.map(hl,subset=["สถานะ"]),use_container_width=True,hide_index=True)
        st.markdown("---")
        st.markdown("#### ⏰ การกระจายวิชาตามช่วงเวลา")
        dist={"เช้า (คาบ 1-3)":0,"บ่าย (คาบ 4-6)":0}
        for cls in st.session_state.classes:
            for d in DAYS:
                for p in PERIODS:
                    slot=st.session_state.schedule[cls][d][p]
                    if slot: dist["เช้า (คาบ 1-3)" if p in MORNING else "บ่าย (คาบ 4-6)"]+=1
        st.bar_chart(pd.DataFrame([dist]).T.rename(columns={0:"จำนวนคาบ"}))
